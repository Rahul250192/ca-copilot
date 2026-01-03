import re
import zipfile
import tempfile
import logging
from pathlib import Path
from typing import Optional, Dict, List

import pdfplumber
from openpyxl import load_workbook

# Template Excel must sit next to this file on disk (Render + local)
import os
TEMPLATE_XLSX = os.path.join(os.path.dirname(__file__), "Statement_sampling.xlsx")

logger = logging.getLogger("extract_batch")
if not logger.handlers:
    logger.setLevel(logging.INFO)


# ---------- PDF TEXT EXTRACTION ----------

def extract_text_first_6_pages(pdf_path: str) -> str:
    """Extract text from the first 6 pages of the given PDF."""
    logger.info(f"  [TEXT] Reading first 6 pages from: {pdf_path}")
    full_text = []
    with pdfplumber.open(pdf_path) as pdf:
        num_pages = min(6, len(pdf.pages))
        logger.info(f"  [TEXT] PDF has {len(pdf.pages)} pages, using first {num_pages}")
        for i in range(num_pages):
            page = pdf.pages[i]
            text = page.extract_text() or ""
            logger.info(f"    [TEXT] Page {i+1} length: {len(text)} chars")
            full_text.append(text)
    return "\n".join(full_text)


# ---------- HELPERS (SHIPPING BILL) ----------

def normalize_sb_date(sb_date: str) -> str:
    """
    Convert SB Date into DD-MM-YYYY format.
    Handles formats:
      - DD-MMM-YY  (e.g., 19-JUN-25)
      - DD/MM/YYYY
      - DD-MM-YYYY
    Returns the input as-is if it doesn't match known formats.
    """
    from datetime import datetime

    sb_date = sb_date.strip()

    # Case 1: DD-MMM-YY  → 19-JUN-25
    try:
        dt = datetime.strptime(sb_date, "%d-%b-%y")
        return dt.strftime("%d-%m-%Y")
    except Exception:
        pass

    # Case 2: DD/MM/YYYY → 20/12/2024
    try:
        dt = datetime.strptime(sb_date, "%d/%m/%Y")
        return dt.strftime("%d-%m-%Y")
    except Exception:
        pass

    # Case 3: DD-MM-YYYY (already ok)
    try:
        dt = datetime.strptime(sb_date, "%d-%m-%Y")
        return dt.strftime("%d-%m-%Y")
    except Exception:
        pass

    # Last: return as-is if unknown pattern
    return sb_date


def parse_invoice_from_lines(lines):
    """
    Generic way to get Invoice No & Date from the '2.INVOICE No. & Dt.' section.

    Works for rows like:
      1 NBRT/06/25-26 15/05/2025
      R 1 RUD/27/24-25 20/12/2024 6390576 CIF
    """
    logger.info("  [INV] Searching for 'INVOICE No' header...")
    header_idx = None
    for i, line in enumerate(lines):
        if "INVOICE No" in line or "Invoice No" in line:
            header_idx = i
            logger.info(f"  [INV] Found invoice header at line {i}: {line!r}")
            break

    if header_idx is None:
        logger.error("  [INV] Could not find any line containing 'INVOICE No'")
        return None, None

    logger.info("  [INV] Scanning rows below invoice header for dd/mm/yyyy...")
    for j in range(header_idx + 1, min(header_idx + 10, len(lines))):
        row = lines[j].strip()
        if not row:
            continue
        logger.info(f"    [INV] Candidate row {j}: {row!r}")
        tokens = row.split()
        for k, tok in enumerate(tokens):
            # dd/mm/yyyy
            if re.fullmatch(r"\d{2}/\d{2}/\d{4}", tok):
                if k == 0:
                    continue
                invoice_date = tok
                invoice_no = tokens[k - 1]
                logger.info(
                    f"  [INV] Parsed invoice_no={invoice_no}, invoice_date={invoice_date}"
                )
                return invoice_no, invoice_date

    logger.error("  [INV] Could not find any dd/mm/yyyy pattern below invoice header")
    return None, None


# ---------- SHIPPING BILL VALUE PARSING ----------

def extract_values_from_text(text: str) -> dict:
    """
    Parse Shipping Bill details from the text (pages 1–6).

    Fields:
        port_code         -> from "Port Code SB No SB Date" value line (last 3 tokens)
        sb_no             -> same line
        sb_date           -> same line, normalized to DD-MM-YYYY
        invoice_no        -> from '2.INVOICE No. & Dt.' block
        invoice_date      -> same as above (dd/mm/yyyy format from PDF)
        invoice_value_inr -> Invoice amount (F/C) * exchange rate
                             (invoice amount from PART-II '1.INVOICE VALUE')
        fob_inr           -> PRIMARY: number immediately before '1.MODE' on the line
                             BACKUP: FOB VALUE (F/C) from PART-II × exchange_rate
    """

    lines = text.splitlines()
    logger.info("  [PARSE] Starting extraction from text")

    invoice_currency: Optional[str] = None
    invoice_value_fc: Optional[float] = None
    fob_inr: Optional[float] = None

    # ----- 1) Port Code / SB No / SB Date -----
    port_code = sb_no = sb_date = None

    logger.info("  [PORT] Searching for 'Port Code SB No SB Date' header...")
    for i, line in enumerate(lines):
        if ("Port Code" in line) and ("SB No" in line) and ("SB Date" in line):
            logger.info(f"  [PORT] Found header at line {i}: {line!r}")
            for j in range(i + 1, min(i + 6, len(lines))):
                next_line = lines[j].strip()
                if not next_line:
                    continue
                logger.info(f"    [PORT] Candidate line {j}: {next_line!r}")
                parts = next_line.split()
                if len(parts) >= 3:
                    raw_sb_date = parts[-1]
                    sb_date = normalize_sb_date(raw_sb_date)
                    port_code = parts[-3]
                    sb_no = parts[-2]
                    logger.info(
                        f"  [PORT] Parsed port_code={port_code}, "
                        f"sb_no={sb_no}, sb_date={sb_date}"
                    )
                    break
            break

    if not (port_code and sb_no and sb_date):
        logger.error("  [PORT] Failed to parse port_code / sb_no / sb_date")
        raise ValueError("Could not find Port Code / SB No / SB Date block")

    # ----- 2) Invoice No & Invoice Date (generic) -----
    invoice_no, invoice_date = parse_invoice_from_lines(lines)
    if not invoice_no or not invoice_date:
        logger.error("  [INV] Final failure: Invoice No & Date not found")
        raise ValueError("Could not find Invoice No & Date")

    # ----- 3) Invoice amount (F/C) & currency from PART-II '1.INVOICE VALUE' -----
    logger.info("  [VAL] Searching PART-II for 'INVOICE VALUE' header...")
    header_idx = None
    for i, line in enumerate(lines):
        norm = line.upper()
        if "INVOICE VALUE" in norm:
            header_idx = i
            logger.info(f"  [VAL] Found 'INVOICE VALUE' header at line {i}: {line!r}")
            break

    if header_idx is None:
        logger.error("  [VAL] Could not find 'INVOICE VALUE' header")
        raise ValueError("Could not find 'INVOICE VALUE' header in text")

    num_pattern = r"\d+(?:,\d{3})*(?:\.\d+)?"

    # Look in following lines for amount + currency
    for j in range(header_idx + 1, min(header_idx + 10, len(lines))):
        row = lines[j].strip()
        if not row:
            continue
        logger.info(f"    [VAL] Candidate PART-II values line {j}: {row!r}")
        tokens = row.split()

        for tok in tokens:
            # Currency: pure 3-letter code like USD, GBP, EUR
            if invoice_currency is None and re.fullmatch(r"[A-Z]{3}", tok):
                invoice_currency = tok
                logger.info(f"    [VAL] Found invoice_currency={invoice_currency}")

            # Numbers: clean letters and punctuation out, keep digits + ',' + '.'
            cleaned = re.sub(r"[^0-9.,]", "", tok)
            if cleaned and re.fullmatch(num_pattern, cleaned):
                value = float(cleaned.replace(",", ""))
                if invoice_value_fc is None:
                    invoice_value_fc = value
                    logger.info(f"    [VAL] Found invoice_value_fc={invoice_value_fc}")

        if invoice_value_fc is not None and invoice_currency is not None:
            logger.info(
                f"  [VAL] Final invoice_value_fc={invoice_value_fc}, "
                f"invoice_currency={invoice_currency}"
            )
            break

    if invoice_value_fc is None:
        logger.error("  [VAL] Could not parse invoice_value_fc from PART-II")
        raise ValueError("Could not parse invoice_value_fc from PART-II")

    if invoice_currency is None:
        logger.warning(
            "  [VAL] Could not find explicit invoice currency near PART-II; "
            "will rely on exchange-rate currency"
        )

    # ----- 4) Exchange rate -----
    logger.info("  [RATE] Searching for '1 <CUR> INR <rate>' for exchange rate...")
    exchange_rate = None

    # Pattern like: "1 USD INR 85.05" or "1 GBP INR 113.15"
    m_rate = re.search(r"1\s+([A-Z]{3})\s+INR\s+([\d.,]+)", text)
    if not m_rate:
        logger.error("  [RATE] Could not find any exchange rate like '1 XXX INR <rate>'")
        raise ValueError("Could not find exchange rate")

    rate_currency = m_rate.group(1)
    exchange_rate = float(m_rate.group(2).replace(",", ""))
    logger.info(
        f"  [RATE] Found exchange_rate={exchange_rate} for currency={rate_currency}"
    )

    if invoice_currency and rate_currency != invoice_currency:
        logger.warning(
            f"  [RATE] Currency mismatch: invoice_currency={invoice_currency}, "
            f"rate_currency={rate_currency}"
        )

    # ----- 5) Compute invoice INR -----
    invoice_value_inr = invoice_value_fc * exchange_rate
    logger.info(
        f"  [VAL] invoice_value_fc={invoice_value_fc}, "
        f"exchange_rate={exchange_rate}, invoice_value_inr={invoice_value_inr}"
    )

    # ----- 6) PRIMARY FOB (INR): number immediately before "1.MODE" -----
    logger.info("  [FOB] Primary: looking for number before '1.MODE'...")
    for line in lines:
        norm = line.upper()
        if "1.MODE" in norm:
            idx = norm.index("1.MODE")
            left = line[:idx]  # everything before '1.MODE'
            logger.info(f"    [FOB] Found '1.MODE' line: {line!r}")
            logger.info(f"    [FOB] Left segment before '1.MODE': {left!r}")

            # Extract all numeric-looking tokens from the left part
            tokens = left.split()
            nums = []
            for tok in tokens:
                cleaned = re.sub(r"[^0-9.,]", "", tok)
                if cleaned and re.fullmatch(num_pattern, cleaned):
                    try:
                        nums.append(float(cleaned.replace(",", "")))
                    except ValueError:
                        continue

            if nums:
                fob_inr = nums[-1]  # last number before '1.MODE' is FOB total
                logger.info(f"  [FOB] PRIMARY FOB (INR) from '... 1.MODE' line = {fob_inr}")
                break

    # ----- 7) BACKUP FOB: from PART-II FOB VALUE (F/C) × exchange_rate -----
    if fob_inr is None:
        logger.info(
            "  [FOB] Primary FOB not found – backup: use PART-II 'FOB VALUE' in F/C × exchange_rate"
        )

        fob_fc_from_part2: Optional[float] = None
        header_idx2 = None
        for i, line in enumerate(lines):
            norm = line.upper()
            if "INVOICE VALUE" in norm and "FOB VALUE" in norm:
                header_idx2 = i
                logger.info(f"    [FOB-BACKUP] Found PART-II header at line {i}: {line!r}")
                break

        if header_idx2 is not None:
            for j in range(header_idx2 + 1, min(header_idx2 + 6, len(lines))):
                row = lines[j].strip()
                if not row:
                    continue
                logger.info(f"    [FOB-BACKUP] Candidate values line {j}: {row!r}")

                tokens = row.split()
                nums = []
                for tok in tokens:
                    cleaned = re.sub(r"[^0-9.,]", "", tok)
                    if cleaned and re.fullmatch(num_pattern, cleaned):
                        try:
                            nums.append(float(cleaned.replace(",", "")))
                        except ValueError:
                            continue

                if len(nums) >= 2:
                    # According to the layout: 1.INVOICE VALUE, 2.FOB VALUE, ...
                    fob_fc_from_part2 = nums[1]
                    logger.info(
                        f"    [FOB-BACKUP] Parsed from PART-II: fob_fc={fob_fc_from_part2}"
                    )
                    fob_inr = fob_fc_from_part2 * exchange_rate
                    logger.info(
                        f"  [FOB-BACKUP] FOB (INR) from PART-II FOB(F/C)×rate = {fob_inr}"
                    )
                    break

        if fob_inr is None:
            logger.error("  [FOB] Could not determine FOB (INR) via primary or backup methods")
            raise ValueError("Could not find FOB INR from summary or PART-II")

    return {
        "port_code": port_code,
        "sb_no": sb_no,
        "sb_date": sb_date,                 # DD-MM-YYYY
        "invoice_no": invoice_no,
        "invoice_date": invoice_date,       # dd/mm/yyyy as in PDF
        "invoice_value_inr": invoice_value_inr,
        "fob_inr": fob_inr,
        "invoice_currency": invoice_currency,
    }


# ---------- BRC PARSING ----------

def extract_brc_info_from_text(text: str) -> dict:
    """
    Extract BRC info from a DGFT eBRC PDF text.

    Returns:
        {
          "sb_no": "6952880",
          "brc_no": "UTIB0000208A00288025",
          "brc_date": "05-03-2025",
          "realised_value": 19745.00,
        }
    """
    # 1) Shipping Bill / Invoice No.
    m = re.search(
        r"Shipping Bill\s*/\s*Invoice No\.\s*([A-Za-z0-9/\-]+)",
        text,
        flags=re.IGNORECASE,
    )
    if not m:
        raise ValueError("Could not find 'Shipping Bill / Invoice No.' in BRC")
    sb_no = m.group(1).strip()

    # 2) Bank Realisation Certificate No + Date
    #    Actual pattern in text:
    #      Bank Realisation Certificate
    #      9 UTIB0000208A00288025 Dated 05-03-2025
    #
    #    We'll first locate the 'Bank Realisation Certificate' block,
    #    then search near it for "<BRC_NO> Dated <dd-mm-yyyy>".
    idx = text.lower().find("bank realisation certificate")
    if idx == -1:
        raise ValueError("Could not find 'Bank Realisation Certificate' in BRC")

    snippet = text[idx: idx + 300]  # small window after that heading

    m = re.search(
        r"\b([A-Z0-9]{8,})\s+Dated\s+([0-9]{2}-[0-9]{2}-[0-9]{4})",
        snippet,
        flags=re.IGNORECASE,
    )
    if not m:
        raise ValueError("Could not find BRC No + 'Dated <date>' pattern near heading")
    brc_no = m.group(1).strip()
    brc_date = m.group(2).strip()

    # 3) Total Realised Value
    #    e.g. "11 Total Realised Value 19,745.00"
    m = re.search(
        r"Total Realised Value\s+([\d,]+\.\d+)",
        text,
        flags=re.IGNORECASE,
    )
    if not m:
        raise ValueError("Could not find 'Total Realised Value' in BRC")
    realised_value = float(m.group(1).replace(",", ""))

    logger.info(
        f"  [BRC] Parsed: sb_no={sb_no}, brc_no={brc_no}, "
        f"brc_date={brc_date}, realised_value={realised_value}"
    )

    return {
        "sb_no": sb_no,
        "brc_no": brc_no,
        "brc_date": brc_date,
        "realised_value": realised_value,
    }


def build_brc_index_from_zip(brc_zip_bytes: Optional[bytes]) -> Dict[str, List[dict]]:
    """
    Build an index:
        sb_no -> [ {brc_no, brc_date, realised_value}, ... ]
    from the BRC ZIP bytes. If brc_zip_bytes is None, returns empty dict.
    """
    if brc_zip_bytes is None:
        return {}

    import io

    logger.info("[BRC] Building BRC index from uploaded ZIP...")
    index: Dict[str, List[dict]] = {}

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        brc_folder = tmp_path / "brc_batch"
        brc_folder.mkdir()

        zip_path = tmp_path / "brc_all.zip"
        with open(zip_path, "wb") as f:
            f.write(brc_zip_bytes)
        logger.info(f"[BRC] Wrote BRC zip to {zip_path}")

        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(brc_folder)
        logger.info(f"[BRC] Extracted BRC ZIP to {brc_folder}")

        brc_files = sorted(
            p for p in brc_folder.rglob("*.pdf") if "__MACOSX" not in str(p)
        )
        logger.info(f"[BRC] Found {len(brc_files)} BRC PDF(s) inside ZIP")

        for brc_path in brc_files:
            logger.info(f"[BRC] Processing BRC file: {brc_path.name}")
            try:
                # BRC is usually 1 page; but just in case, read first 2 pages
                with pdfplumber.open(str(brc_path)) as pdf:
                    text_parts = []
                    for page in pdf.pages[:2]:
                        text_parts.append(page.extract_text() or "")
                    text = "\n".join(text_parts)

                info = extract_brc_info_from_text(text)
                sb_no = info["sb_no"]
                index.setdefault(sb_no, []).append(info)
            except Exception as e:
                logger.exception(f"[BRC] ERROR processing {brc_path.name}: {e}")

    logger.info(f"[BRC] Built BRC index for {len(index)} SB numbers")
    return index


# ---------- EXCEL APPEND ----------

def append_to_excel_row(ws, row_idx: int,
                        shipping_values: Optional[dict],
                        brc_values: Optional[dict]):
    """
    Append a single data row to Statement 3 sheet.

    If shipping_values is not None, writes the "Document Details" side
    (columns 1–10). If shipping_values is None, leaves those blank.

    BRC/FIRC columns (11–13) are filled if brc_values is provided:
        11: BRC/FIRC No.
        12: BRC/FIRC Date (dd-mm-yyyy)
        13: BRC/FIRC Value
    """

    # Document side
    if shipping_values is not None:
        sr_no = row_idx - 11  # row 12 -> Sr. No. 1
        logger.info(
            f"  [EXCEL] Writing SHIPPING row {row_idx} (Sr.No={sr_no}) "
            f"for invoice {shipping_values['invoice_no']}"
        )
        ws.cell(row=row_idx, column=1, value=sr_no)                               # Sr. No.
        ws.cell(row=row_idx, column=2, value="Invoice")                           # Type of Document
        ws.cell(row=row_idx, column=3, value=shipping_values["invoice_no"])       # Invoice No
        ws.cell(row=row_idx, column=4, value=shipping_values["invoice_date"])     # Invoice Date
        ws.cell(row=row_idx, column=5, value=shipping_values["invoice_value_inr"])# Value (INR)
        ws.cell(row=row_idx, column=6, value="G")                                 # G/S
        ws.cell(row=row_idx, column=7, value=shipping_values["port_code"])        # Port Code
        ws.cell(row=row_idx, column=8, value=shipping_values["sb_no"])            # SB No
        ws.cell(row=row_idx, column=9, value=shipping_values["sb_date"])          # SB Date
        ws.cell(row=row_idx, column=10, value=shipping_values["fob_inr"])         # FOB (INR)
    else:
        logger.info(f"  [EXCEL] Writing BRC-only continuation row {row_idx}")
        # Leave columns 1–10 blank

    # BRC/FIRC side (11–13)
    if brc_values is not None:
        ws.cell(row=row_idx, column=13, value=brc_values["brc_no"])       # BRC/FIRC No.
        ws.cell(row=row_idx, column=14, value=brc_values["brc_date"])     # Date (dd-mm-yyyy)
        ws.cell(row=row_idx, column=15, value=brc_values["realised_value"])  # Value
    # else: leave BRC columns blank


def append_shipping_and_brc_rows(ws,
                                 start_row: int,
                                 shipping_values: dict,
                                 brc_list: List[dict]) -> int:
    """
    For a single shipping bill:

    - If brc_list is empty: one row with shipping only.
    - If brc_list has N items:
        - First row: shipping + first BRC
        - Next (N-1) rows: only BRC columns, document side empty.

    Returns: next free row index after writing.
    """
    if not brc_list:
        append_to_excel_row(ws, start_row, shipping_values, None)
        return start_row + 1

    # First row: shipping + first BRC
    append_to_excel_row(ws, start_row, shipping_values, brc_list[0])
    current_row = start_row + 1

    # Remaining BRCs: only BRC columns
    for brc in brc_list[1:]:
        append_to_excel_row(ws, current_row, None, brc)
        current_row += 1

    return current_row


# ---------- CORE FOR WEB API ----------

def process_zip_bytes(shipping_zip_bytes: bytes,
                      brc_zip_bytes: Optional[bytes] = None) -> bytes:
    """
    Takes:
        - shipping_zip_bytes: ZIP bytes containing N Shipping Bill PDFs
        - brc_zip_bytes: ZIP bytes containing BRC PDFs (optional)

    Runs extraction logic and returns Excel file bytes.
    """
    import io

    logger.info("[ZIP] Starting process_zip_bytes(...)")

    # Build BRC index (sb_no -> list of brc dicts)
    brc_index = build_brc_index_from_zip(brc_zip_bytes)

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        pdf_folder = tmp_path / "pdf_batch"
        pdf_folder.mkdir()

        zip_path = tmp_path / "all_pdfs.zip"
        with open(zip_path, "wb") as f:
            f.write(shipping_zip_bytes)
        logger.info(f"[ZIP] Wrote uploaded shipping zip to {zip_path}")

        # Extract Shipping Bill PDFs
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(pdf_folder)
        logger.info(f"[ZIP] Extracted Shipping ZIP to {pdf_folder}")

        pdf_files = sorted(
            p for p in pdf_folder.rglob("*.pdf") if "__MACOSX" not in str(p)
        )
        logger.info(f"[ZIP] Found {len(pdf_files)} Shipping Bill PDF(s) inside ZIP")

        if not pdf_files:
            raise ValueError("No PDF files found in uploaded Shipping ZIP")

        # Load template Excel from repo
        template_path = Path(TEMPLATE_XLSX)
        if not template_path.exists():
            logger.error(f"[EXCEL] Template not found at {template_path}")
            raise FileNotFoundError(f"Template not found: {TEMPLATE_XLSX}")
        logger.info(f"[EXCEL] Loading template workbook from {template_path}")
        wb = load_workbook(template_path)
        ws = wb["Statement 3"]
        current_row = 12  # first data row

        for pdf_path in pdf_files:
            logger.info(f"[PDF] Processing Shipping Bill file: {pdf_path.name}")
            try:
                text = extract_text_first_6_pages(str(pdf_path))
                values = extract_values_from_text(text)
                logger.info(f"[PDF] Extracted shipping values: {values}")

                sb_no = values["sb_no"]
                brcs_for_sb = brc_index.get(sb_no, [])
                logger.info(
                    f"[PDF] Found {len(brcs_for_sb)} BRC record(s) for SB No {sb_no}"
                )

                current_row = append_shipping_and_brc_rows(
                    ws, current_row, values, brcs_for_sb
                )
            except Exception as e:
                logger.exception(
                    f"[PDF] ERROR processing {pdf_path.name}: {e}"
                )

        # Save Excel into memory
        out_mem = io.BytesIO()
        wb.save(out_mem)
        out_mem.seek(0)
        logger.info("[EXCEL] Finished writing workbook to memory")
        return out_mem.read()


# ---------- Optional CLI: local testing ----------

def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    )
    SHIPPING_ZIP = "all_pdfs.zip"
    BRC_ZIP = "all_brc.zip"  # optional

    if not Path(SHIPPING_ZIP).exists():
        raise SystemExit(f"Shipping ZIP file not found: {SHIPPING_ZIP}")

    shipping_bytes = Path(SHIPPING_ZIP).read_bytes()
    brc_bytes = Path(BRC_ZIP).read_bytes() if Path(BRC_ZIP).exists() else None

    logger.info(f"[CLI] Running local test on {SHIPPING_ZIP} (BRC: {BRC_ZIP})")
    excel_bytes = process_zip_bytes(shipping_bytes, brc_bytes)
    out_path = Path("Statement_output_filled.xlsx")
    out_path.write_bytes(excel_bytes)
    logger.info(f"[CLI] Saved {out_path}")


if __name__ == "__main__":
    main()
