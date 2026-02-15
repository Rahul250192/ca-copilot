
import re
import pdfplumber
import logging
from typing import Optional, Dict, List, Tuple
import os
import shutil
import tempfile
import zipfile
import glob
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO

# Tesseract check
try:
    import pytesseract
    from PIL import Image
    HAS_PYTESSERACT = True
except ImportError:
    HAS_PYTESSERACT = False

logger = logging.getLogger("extract_firc_details")

def detect_bank(text: str) -> Optional[str]:
    """Detects if the FIRC is from ICICI or HDFC based on keywords."""
    if not text:
        return None
    text_upper = text.upper()
    if "ICICI BANK" in text_upper:
        return "ICICI"
    elif "HDFC BANK" in text_upper:
        return "HDFC"
    return None

def normalize_date(date_str: str) -> str:
    """Normalizes date to DD-MM-YYYY format."""
    if not date_str:
        return None
    date_str = date_str.strip()
    
    formats = [
        "%d-%b-%y",       # 06-Oct-25
        "%b %d, %Y",       # Dec 04, 2024
        "%d-%b-%Y",       # 18-Oct-2024
        "%d-%m-%Y",       # 04-12-2024
        "%d/%m/%Y",       # 04/12/2024
        "%Y-%m-%d"        # 2024-12-04
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            pass
    return date_str

def _parse_icici(text: str) -> Dict[str, Optional[str]]:
    """Parses ICICI FIRC text - Updated."""
    logger.info("Parsing as ICICI FIRC")
    data = {"bank": "ICICI", "firc_no": None, "firc_date": None, "realised_value": None, "currency": "INR"}
    
    # 1. Reference Number
    m_ref = re.search(r"Reference\s+No\s*[:\.]?\s*([A-Za-z0-9]+)", text, re.IGNORECASE)
    if m_ref:
        data["firc_no"] = m_ref.group(1).strip()
        
    # 2. Date
    m_date = re.search(r"Dated\s*[:\.]?\s*([A-Za-z]{3}\s+\d{2},\s+\d{4})", text, re.IGNORECASE)
    if m_date:
        data["firc_date"] = normalize_date(m_date.group(1).strip())
    else:
        m_date = re.search(r"\bDate\s*[:\.]?\s*([A-Za-z]{3}\s+\d{2},\s+\d{4})", text, re.IGNORECASE)
        if m_date:
            data["firc_date"] = normalize_date(m_date.group(1).strip())

    # 3. Amount (INR)
    m_amt = re.search(r"INR\s+([\d,]+\.\d{2})", text)
    if m_amt:
        data["realised_value"] = m_amt.group(1).replace(",", "")
        
    return data

def _parse_hdfc(text: str) -> Dict[str, Optional[str]]:
    """Parses HDFC FIRC text - Updated."""
    logger.info("Parsing as HDFC FIRC")
    data = {"bank": "HDFC", "firc_no": None, "firc_date": None, "realised_value": None, "currency": "INR"}

    # 1. Reference Number (Inward No)
    lines = text.splitlines()
    inward_idx = -1
    for i, line in enumerate(lines):
        if "Inward No" in line and "Sender Ref No" in line:
            inward_idx = i
            break
            
    if inward_idx != -1 and inward_idx + 1 < len(lines):
        val_line = lines[inward_idx + 1].strip()
        tokens = val_line.split()
        if tokens:
            data["firc_no"] = tokens[0] 
            raw_date = None
            if len(tokens) >= 1:
                raw_date = tokens[-1]
            if raw_date:
                data["firc_date"] = normalize_date(raw_date)

    # 3. Amount (INR)
    m_amt = re.search(r"(?:USD|EUR|GBP|AUD|CAD|SGD)\s+[\d,]+\.\d{2}\s+[\d\.]+\s+([\d,]+\.\d{2})", text)
    if m_amt:
        data["realised_value"] = m_amt.group(1).replace(",", "")

    return data

def extract_firc_text(file_path: str) -> str:
    """Extracts text from PDF or Image (JPG/JPEG)."""
    ext = os.path.splitext(file_path)[1].lower()
    text = ""
    
    if ext == ".pdf":
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages[:2]: # Max 2 pages
                    text += page.extract_text() or ""
                    text += "\n"
        except Exception as e:
            logger.error(f"Error reading PDF {file_path}: {e}")
            
    elif ext in [".jpg", ".jpeg", ".png"]:
        if HAS_PYTESSERACT:
            try:
                pytesseract.get_tesseract_version()
                text = pytesseract.image_to_string(Image.open(file_path))
            except Exception as e:
                logger.warning(f"OCR Failed for {file_path}. Is Tesseract installed? Error: {e}")
        else:
            logger.warning(f"Skipping Image {file_path}: pytesseract not installed.")
            
    return text

def extract_firc_data(file_path: str) -> Dict[str, Optional[str]]:
    """Extracts data from FIRC file (PDF or Image)."""
    text = extract_firc_text(file_path)
    bank = detect_bank(text)
    
    if bank == "ICICI":
        return _parse_icici(text)
    elif bank == "HDFC":
        return _parse_hdfc(text)
    else:
        logger.warning(f"Could not detect supported bank for {file_path}")
        return {"bank": "UNKNOWN", "error": "Bank not detected"}

def extract_invoice_data(pdf_path: str) -> Dict[str, Optional[str]]:
    """
    Parses Invoice PDF to extract fields.
    """
    data = {
        "invoice_no": None, "invoice_date": None, "invoice_val": None,
        "sb_no": None, "sb_date": None, "port_code": None
    }
    
    try:
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:1]: 
                text += page.extract_text() or ""
        
        # Regex Strategies based on actual dump:
        # Invoice no: 25-26-033
        m_inv = re.search(r"Invoice\s*no\s*[:\-\.]?\s*([A-Za-z0-9/\-]+)", text, re.IGNORECASE)
        if m_inv:
            data["invoice_no"] = m_inv.group(1).strip()
            
        # Invoice date: 06-Oct-25
        m_date = re.search(r"Invoice\s*date\s*[:\-\.]?\s*([0-9A-Za-z\-]+)", text, re.IGNORECASE)
        if m_date:
            data["invoice_date"] = normalize_date(m_date.group(1).strip())
            
        # SB No: "SB No[:.] <val>"
        m_sb = re.search(r"S\.?B\.?\s*No\.?\s*[:\-\.]?\s*(\d+)", text, re.IGNORECASE)
        if m_sb:
            data["sb_no"] = m_sb.group(1).strip()
            
        # SB Date
        m_sb_date = re.search(r"S\.?B\.?\s*Date\s*[:\-\.]?\s*(\d{2}[-/]\d{2}[-/]\d{4})", text, re.IGNORECASE)
        if m_sb_date:
            data["sb_date"] = normalize_date(m_sb_date.group(1).strip())
            
        # Port Code
        m_port = re.search(r"Port\s*Code\s*[:\-\.]?\s*([A-Z]{5,})", text, re.IGNORECASE)
        if m_port:
            data["port_code"] = m_port.group(1).strip()
            
        # Invoice Value (Total)
        # Format: TOTAL $ 5,750.00 USD
        m_val = re.search(r"TOTAL\s*[\$€£]?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)
        if not m_val:
             m_val = re.search(r"(?:Total|Grand\s+Total|Amount)\s*[:\-\.]?\s*(?:[\$€£A-Z]{1,3})?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)

        if m_val:
            data["invoice_val"] = m_val.group(1).replace(",", "")
            
    except Exception as e:
        logger.error(f"Error parsing Invoice {pdf_path}: {e}")
        
    return data

def process_statement3_workflow(invoice_zip_bytes: bytes, firc_zip_bytes: bytes) -> bytes:
    """
    Processes Invoice ZIP and FIRC ZIP.
    Maps them 1:1 by SORTED filename order.
    Generates Statement 3 Excel.
    Returns: Excel bytes.
    """
    
    with tempfile.TemporaryDirectory() as temp_dir:
        inv_dir = os.path.join(temp_dir, "invoices")
        firc_dir = os.path.join(temp_dir, "fircs")
        os.makedirs(inv_dir, exist_ok=True)
        os.makedirs(firc_dir, exist_ok=True)
        
        # 1. Extract Invoices
        with zipfile.ZipFile(BytesIO(invoice_zip_bytes)) as z:
            z.extractall(inv_dir)
            
        # 2. Extract FIRCs
        with zipfile.ZipFile(BytesIO(firc_zip_bytes)) as z:
            z.extractall(firc_dir)
            
        # 3. Get Sorted File Lists
        inv_files = sorted([
            os.path.join(dp, f) for dp, dn, filenames in os.walk(inv_dir) for f in filenames 
            if f.lower().endswith('.pdf')
        ])
        
        firc_files = sorted([
            os.path.join(dp, f) for dp, dn, filenames in os.walk(firc_dir) for f in filenames
            if f.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png'))
        ])
        
        logger.info(f"Found {len(inv_files)} Invoices and {len(firc_files)} FIRCs")
        
        # 4. Prepare Excel
        TEMPLATE_XLSX = os.path.join(os.path.dirname(__file__), "Statement_sampling.xlsx")
        wb = load_workbook(TEMPLATE_XLSX)
        ws = wb["Statement 3"]
        current_row = 12
        sr_no = 1
        
        pairs = list(zip(inv_files, firc_files))
        
        for inv_path, firc_path in pairs:
            logger.info(f"Processing Pair: {os.path.basename(inv_path)} <-> {os.path.basename(firc_path)}")
            
            # Extract Data
            inv_data = extract_invoice_data(inv_path)
            firc_data = extract_firc_data(firc_path)
            
            # Write to Excel
            ws.cell(row=current_row, column=1, value=sr_no)
            
            # -- Document Details --
            # Col 2: Type of document
            ws.cell(row=current_row, column=2, value="INVOICE")
            
            # Col 3: Invoice No
            if inv_data.get("invoice_no"):
                 ws.cell(row=current_row, column=3, value=inv_data["invoice_no"])

            # Col 4: Invoice Date
            if inv_data.get("invoice_date"):
                 ws.cell(row=current_row, column=4, value=inv_data["invoice_date"])
                 
            # Col 5: Invoice Value
            if inv_data.get("invoice_val"):
                try:
                    ws.cell(row=current_row, column=5, value=float(inv_data["invoice_val"]))
                except:
                    ws.cell(row=current_row, column=5, value=inv_data["invoice_val"])
            
            # Col 6: Goods/Services (G/S) -> "S"
            ws.cell(row=current_row, column=6, value="S")

            # -- Shipping Bill Details --
            # Col 7: Port Code
            if inv_data.get("port_code"):
                ws.cell(row=current_row, column=7, value=inv_data["port_code"])
                
            # Col 8: SB No
            if inv_data.get("sb_no"):
                ws.cell(row=current_row, column=8, value=inv_data["sb_no"])
                
            # Col 9: SB Date
            if inv_data.get("sb_date"):
                ws.cell(row=current_row, column=9, value=inv_data["sb_date"])
            
            # -- FIRC Details (Cols 13-15) --
            # Col 13: BRC/FIRC No
            if firc_data.get("firc_no"):
                ws.cell(row=current_row, column=13, value=firc_data["firc_no"])
            # Col 14: Date
            if firc_data.get("firc_date"):
                ws.cell(row=current_row, column=14, value=firc_data["firc_date"])
            # Col 15: Value (Numeric)
            if firc_data.get("realised_value"):
                val_str = firc_data["realised_value"]
                try:
                    ws.cell(row=current_row, column=15, value=float(val_str))
                except:
                    ws.cell(row=current_row, column=15, value=val_str)
                    
            current_row += 1
            sr_no += 1
            
        # Return Excel Bytes
        out_buffer = BytesIO()
        wb.save(out_buffer)
        out_buffer.seek(0)
        return out_buffer.read()
