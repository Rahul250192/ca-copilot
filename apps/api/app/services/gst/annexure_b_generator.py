# annexure_b_generator.py
import io
import os
from typing import Dict, Any, List, Sequence, Union

from openpyxl import load_workbook


# ---------------- Local filenames (must be in same folder as this file on server) ----------------
MASTER_FILENAME = "MASTER_SHEET.xlsx"
TEMPLATE_FILENAME = "Annexure_b_template.xlsx"

# ---------------- GSTR2B B2B ----------------
GSTR2B_B2B_SHEET_NAME = "B2B"
GSTR2B_HEADER_ROW_MAIN = 5
GSTR2B_HEADER_ROW_SUB = 6
GSTR2B_DATA_START_ROW = 7

# ---------------- Template ----------------
TEMPLATE_SHEET_NAME = "Sheet1"
TEMPLATE_DATA_START_ROW = 5  # rows 3-4 headers, data starts from row 5

# ---------------- Master ----------------
MASTER_SHEET_NAME = "Sheet1"
MASTER_HEADER_ROW = 1


# ---------------- Helpers ----------------
def _norm(x: Any) -> str:
    return str(x).strip().lower() if x not in (None, "") else ""


def _get_required_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    return wb[sheet_name]


def _build_gstr2b_b2b_headers(ws) -> Dict[str, int]:
    """
    Headers split across row 5 & 6.
    Effective header = row6 if present else row5
    """
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h5 = _norm(ws.cell(GSTR2B_HEADER_ROW_MAIN, c).value)
        h6 = _norm(ws.cell(GSTR2B_HEADER_ROW_SUB, c).value)
        final = h6 if h6 else h5
        if final:
            headers[final] = c
    return headers


def _build_single_row_headers(ws, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm(ws.cell(header_row, c).value)
        if key:
            headers[key] = c
    return headers


def _find_first_empty_row(ws, start_row: int) -> int:
    r = start_row
    while True:
        if ws.cell(r, 1).value in (None, ""):
            return r
        r += 1


def _safe_float(x: Any) -> float:
    if x in (None, ""):
        return 0.0
    try:
        return float(x)
    except Exception:
        return 0.0


def _key_gstin(gstin: Any) -> str:
    return str(gstin).strip()


def _load_local_workbook(path: str, data_only: bool):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Required file not found on server: {path}")
    return load_workbook(path, data_only=data_only)


def _validate_required_b2b_headers(b2b_headers: Dict[str, int]) -> None:
    required = [
        "gstin of supplier",
        "trade/legal name",
        "invoice number",
        "invoice date",
        "invoice value(₹)",
        "central tax(₹)",
        "state/ut tax(₹)",
        "integrated tax(₹)",
        "cess(₹)",
    ]
    missing = [x for x in required if x not in b2b_headers]
    if missing:
        raise ValueError(
            f"GSTR2B B2B missing required columns: {missing}. Found: {sorted(b2b_headers.keys())}"
        )


# ---------------- Master lookup (GSTIN-only) ----------------
def build_master_lookup(master_ws) -> Dict[str, Dict[str, Any]]:
    """
    MASTER_SHEET.xlsx contains at least:
      - GSTIN of the supplier
      - HSN/SAC
      - Inputs/ Input Services/ capital goods

    Build mapping:
      GSTIN -> {hsn, category}

    If multiple rows exist per GSTIN:
      first non-empty HSN/category wins (minimal logic).
    """
    headers = _build_single_row_headers(master_ws, MASTER_HEADER_ROW)

    gstin_col = headers.get(_norm("GSTIN of the supplier"))
    hsn_col = headers.get(_norm("HSN/SAC"))

    category_col = headers.get(_norm("Inputs/ Input Services/ capital goods"))
    if not category_col:
        # handle common line-break variant
        category_col = headers.get(_norm("Inputs/ Input\nServices/ cap\nital goods"))

    if not gstin_col:
        raise ValueError(f"MASTER missing 'GSTIN of the supplier'. Found: {sorted(headers.keys())}")
    if not hsn_col:
        raise ValueError(f"MASTER missing 'HSN/SAC'. Found: {sorted(headers.keys())}")
    if not category_col:
        raise ValueError(
            "MASTER missing 'Inputs/ Input Services/ capital goods'. "
            f"Found: {sorted(headers.keys())}"
        )

    lookup: Dict[str, Dict[str, Any]] = {}

    for r in range(MASTER_HEADER_ROW + 1, master_ws.max_row + 1):
        gstin = master_ws.cell(r, gstin_col).value
        if gstin in (None, ""):
            continue

        k = _key_gstin(gstin)
        hsn = master_ws.cell(r, hsn_col).value
        category = master_ws.cell(r, category_col).value

        if k not in lookup:
            lookup[k] = {"hsn": hsn, "category": category}
        else:
            # keep first non-empty values (minimal, predictable)
            if lookup[k].get("hsn") in (None, "") and hsn not in (None, ""):
                lookup[k]["hsn"] = hsn
            if lookup[k].get("category") in (None, "") and category not in (None, ""):
                lookup[k]["category"] = category

    return lookup


def _write_one_b2b_row_into_template(
    *,
    out_ws,
    out_row: int,
    s_no: int,
    gstin: Any,
    supplier: Any,
    inv_no: Any,
    inv_date: Any,
    inv_value: Any,
    cgst: Any,
    sgst: Any,
    igst: Any,
    cess: Any,
    master_lookup: Dict[str, Dict[str, Any]],
) -> float:
    """
    Writes one logical row into output sheet.
    Returns eligible ITC amount (float) (mostly for debugging/extension).
    """
    m = master_lookup.get(_key_gstin(gstin), {})
    hsn = m.get("hsn")
    category = (m.get("category") or "").strip()

    # Eligible For ITC rule:
    # - Yes if Inputs or Input Services
    # - No if Capital goods
    cat_norm = category.lower()
    eligible_itc_yes = ("capital" not in cat_norm)

    eligible_itc_amount = 0.0
    if eligible_itc_yes:
        eligible_itc_amount = (
            _safe_float(cgst) + _safe_float(sgst) + _safe_float(igst) + _safe_float(cess)
        )

    # Write into template columns A..N
    out_ws.cell(out_row, 1).value = s_no                   # A S.No
    out_ws.cell(out_row, 2).value = _key_gstin(gstin)      # B GSTIN
    out_ws.cell(out_row, 3).value = supplier               # C Name of The Supplier
    out_ws.cell(out_row, 4).value = inv_no                 # D Invoice No
    out_ws.cell(out_row, 5).value = inv_date               # E Date
    out_ws.cell(out_row, 6).value = inv_value              # F Value

    out_ws.cell(out_row, 7).value = category or "Inputs"   # G Inputs / Input Services / Capital Goods
    out_ws.cell(out_row, 8).value = hsn                    # H HSN/SAC

    out_ws.cell(out_row, 9).value = cgst                   # I Central Tax
    out_ws.cell(out_row, 10).value = sgst                  # J State Tax
    out_ws.cell(out_row, 11).value = igst                  # K Integrated Tax
    out_ws.cell(out_row, 12).value = cess                  # L Cess

    out_ws.cell(out_row, 13).value = "Yes" if eligible_itc_yes else "No"  # M Eligible For ITC
    out_ws.cell(out_row, 14).value = eligible_itc_amount                  # N Amount of Eligible ITC

    return eligible_itc_amount


# ---------------- Unified generator: accepts single OR multiple files ----------------
def generate_annexure_b(
    gstr2b_excel: Union[bytes, Sequence[bytes]],
    base_dir: str,
) -> bytes:
    """
    Input:
      - single GSTR2B file bytes, OR
      - multiple GSTR2B file bytes (list/tuple/etc.)

    Reads MASTER_SHEET.xlsx + Annexure_b_template.xlsx from base_dir
    Output: Annexure_b.xlsx bytes containing merged rows from all provided files (append merge).
    """
    # Normalize input to list[bytes]
    if isinstance(gstr2b_excel, (bytes, bytearray)):
        gstr2b_files: List[bytes] = [bytes(gstr2b_excel)]
    else:
        gstr2b_files = [bytes(x) for x in gstr2b_excel]

    if not gstr2b_files:
        raise ValueError("No GSTR2B files provided")

    master_path = os.path.join(base_dir, MASTER_FILENAME)
    template_path = os.path.join(base_dir, TEMPLATE_FILENAME)

    master_wb = _load_local_workbook(master_path, data_only=True)
    template_wb = _load_local_workbook(template_path, data_only=False)

    master_ws = _get_required_sheet(master_wb, MASTER_SHEET_NAME)
    out_ws = _get_required_sheet(template_wb, TEMPLATE_SHEET_NAME)

    # Master lookup once
    master_lookup = build_master_lookup(master_ws)

    # Output cursor once (we append all files into same sheet)
    out_row = _find_first_empty_row(out_ws, TEMPLATE_DATA_START_ROW)
    s_no = 1

    for file_bytes in gstr2b_files:
        gstr_wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        b2b_ws = _get_required_sheet(gstr_wb, GSTR2B_B2B_SHEET_NAME)

        b2b_headers = _build_gstr2b_b2b_headers(b2b_ws)
        _validate_required_b2b_headers(b2b_headers)

        for r in range(GSTR2B_DATA_START_ROW, b2b_ws.max_row + 1):
            gstin = b2b_ws.cell(r, b2b_headers["gstin of supplier"]).value
            if gstin in (None, ""):
                continue

            supplier = b2b_ws.cell(r, b2b_headers["trade/legal name"]).value
            inv_no = b2b_ws.cell(r, b2b_headers["invoice number"]).value
            inv_date = b2b_ws.cell(r, b2b_headers["invoice date"]).value
            inv_value = b2b_ws.cell(r, b2b_headers["invoice value(₹)"]).value

            cgst = b2b_ws.cell(r, b2b_headers["central tax(₹)"]).value
            sgst = b2b_ws.cell(r, b2b_headers["state/ut tax(₹)"]).value
            igst = b2b_ws.cell(r, b2b_headers["integrated tax(₹)"]).value
            cess = b2b_ws.cell(r, b2b_headers["cess(₹)"]).value

            _write_one_b2b_row_into_template(
                out_ws=out_ws,
                out_row=out_row,
                s_no=s_no,
                gstin=gstin,
                supplier=supplier,
                inv_no=inv_no,
                inv_date=inv_date,
                inv_value=inv_value,
                cgst=cgst,
                sgst=sgst,
                igst=igst,
                cess=cess,
                master_lookup=master_lookup,
            )

            out_row += 1
            s_no += 1

    out = io.BytesIO()
    template_wb.save(out)
    return out.getvalue()
