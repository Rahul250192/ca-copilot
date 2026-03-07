"""
HSN/SAC Locator API
Loads HSN_SAC.xlsx into memory for fast fuzzy description matching.
"""
import os
import io
from typing import List, Optional
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

import openpyxl
from thefuzz import fuzz

from app.api import deps
from app.models.models import User

router = APIRouter()

# ─── In-memory HSN/SAC data ───
HSN_DATA: List[dict] = []
SAC_DATA: List[dict] = []
_loaded = False


def _load_data():
    global HSN_DATA, SAC_DATA, _loaded
    if _loaded:
        return

    xlsx_path = os.path.join(os.path.dirname(__file__), "../../services/gst/HSN_SAC.xlsx")
    xlsx_path = os.path.abspath(xlsx_path)

    if not os.path.exists(xlsx_path):
        print(f"⚠️ HSN_SAC.xlsx not found at {xlsx_path}")
        _loaded = True
        return

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Load HSN_MSTR
    if "HSN_MSTR" in wb.sheetnames:
        ws = wb["HSN_MSTR"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                HSN_DATA.append({
                    "code": str(row[0]).strip(),
                    "description": str(row[1]).strip(),
                    "type": "HSN",
                })

    # Load SAC_MSTR
    if "SAC_MSTR" in wb.sheetnames:
        ws = wb["SAC_MSTR"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                SAC_DATA.append({
                    "code": str(row[0]).strip(),
                    "description": str(row[1]).strip(),
                    "type": "SAC",
                })

    wb.close()
    _loaded = True
    print(f"✅ HSN Locator: Loaded {len(HSN_DATA)} HSN + {len(SAC_DATA)} SAC codes")


# ─── Schemas ───
class HSNMatch(BaseModel):
    code: str
    description: str
    type: str  # HSN or SAC
    score: int  # match score 0-100


class SingleLookupRequest(BaseModel):
    description: str
    limit: int = 10
    include_sac: bool = True


class BulkItem(BaseModel):
    description: str
    matched_code: Optional[str] = None
    matched_description: Optional[str] = None
    match_type: Optional[str] = None
    score: Optional[int] = None


# ─── Search Function ───
def search_hsn(query: str, data: List[dict], limit: int = 10, min_score: int = 40) -> List[dict]:
    """Fuzzy search HSN/SAC by description."""
    query_upper = query.upper().strip()

    # Exact code match first
    for item in data:
        if item["code"] == query_upper:
            return [{"code": item["code"], "description": item["description"], "type": item["type"], "score": 100}]

    # Fuzzy description matching
    scored = []
    for item in data:
        # Use token_set_ratio for best matching with partial/reordered tokens
        score = fuzz.token_set_ratio(query_upper, item["description"].upper())
        if score >= min_score:
            scored.append({
                "code": item["code"],
                "description": item["description"],
                "type": item["type"],
                "score": score,
            })

    # Sort by score descending, then by code length (shorter = more general category)
    scored.sort(key=lambda x: (-x["score"], len(x["code"])))
    return scored[:limit]


# ─── Endpoints ───

@router.get("/search")
async def search_hsn_code(
    q: str,
    limit: int = 10,
    include_sac: bool = True,
    current_user: User = Depends(deps.get_current_user),
) -> List[HSNMatch]:
    """Search HSN/SAC codes by description or code number."""
    _load_data()
    if not q or len(q.strip()) < 2:
        raise HTTPException(status_code=400, detail="Query must be at least 2 characters")

    all_data = HSN_DATA + (SAC_DATA if include_sac else [])
    results = search_hsn(q, all_data, limit=limit)
    return results


@router.post("/bulk-validate")
async def bulk_validate(
    file: UploadFile = File(...),
    current_user: User = Depends(deps.get_current_user),
):
    """
    Upload Excel with HSN_Description column.
    Returns the same Excel with matched HSN_Code, HSN_Matched_Description, and Match_Score columns.
    """
    _load_data()

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Find the HSN_Description column
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    desc_col = None
    for i, h in enumerate(headers):
        if h.upper() in ("HSN_DESCRIPTION", "HSN DESCRIPTION", "DESCRIPTION", "PRODUCT", "PRODUCT_NAME", "ITEM", "ITEM_NAME", "ITEM_DESCRIPTION"):
            desc_col = i
            break

    if desc_col is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find 'HSN_Description' column. Expected columns: HSN_Description, Description, Product, Item"
        )

    # Add output columns
    out_headers = ["HSN_Code", "Matched_Description", "Match_Type", "Match_Score"]
    next_col = ws.max_column + 1
    for i, h in enumerate(out_headers):
        ws.cell(row=1, column=next_col + i, value=h)

    all_data = HSN_DATA + SAC_DATA

    # Process each row
    for row_idx in range(2, ws.max_row + 1):
        desc_value = ws.cell(row=row_idx, column=desc_col + 1).value
        if not desc_value or str(desc_value).strip() == "":
            continue

        matches = search_hsn(str(desc_value), all_data, limit=1)
        if matches:
            best = matches[0]
            ws.cell(row=row_idx, column=next_col, value=best["code"])
            ws.cell(row=row_idx, column=next_col + 1, value=best["description"])
            ws.cell(row=row_idx, column=next_col + 2, value=best["type"])
            ws.cell(row=row_idx, column=next_col + 3, value=best["score"])
        else:
            ws.cell(row=row_idx, column=next_col, value="NOT_FOUND")
            ws.cell(row=row_idx, column=next_col + 3, value=0)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=HSN_Validated_{file.filename}"}
    )
